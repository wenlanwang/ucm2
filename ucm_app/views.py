from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.db.models import Q
from django.utils import timezone
from django.http import HttpResponse
import json
import xlrd
import zipfile
import io
from datetime import datetime, time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import (
    ManufacturerVersionInfo, ColumnOptions, UCMDeviceInventory,
    UCMRequirement, TemplateConfig, UCMDateConfig
)
from .serializers import (
    UserSerializer, ManufacturerVersionInfoSerializer, ColumnOptionsSerializer,
    UCMDeviceInventorySerializer, UCMRequirementSerializer, TemplateConfigSerializer
)


class ManufacturerVersionInfoViewSet(viewsets.ModelViewSet):
    """厂商版本信息管理API"""
    queryset = ManufacturerVersionInfo.objects.all()
    serializer_class = ManufacturerVersionInfoSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = ManufacturerVersionInfo.objects.all()
        device_type = self.request.query_params.get('device_type')
        if device_type:
            queryset = queryset.filter(device_type=device_type)
        return queryset
    
    @action(detail=False, methods=['get'])
    def get_manufacturers(self, request):
        """根据设备类型获取品牌(厂商)列表"""
        device_type = request.query_params.get('device_type')
        if not device_type:
            return Response({'error': 'device_type参数必填'}, status=status.HTTP_400_BAD_REQUEST)
        
        manufacturers = ManufacturerVersionInfo.objects.filter(
            device_type=device_type
        ).values_list('manufacturer', flat=True).distinct()
        return Response(list(manufacturers))
    
    @action(detail=False, methods=['get'])
    def get_versions(self, request):
        """根据设备类型和品牌(厂商)获取版本列表"""
        device_type = request.query_params.get('device_type')
        manufacturer = request.query_params.get('manufacturer')
        if not device_type or not manufacturer:
            return Response({'error': 'device_type和manufacturer参数必填'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        versions = ManufacturerVersionInfo.objects.filter(
            device_type=device_type,
            manufacturer=manufacturer
        ).values_list('version', flat=True).distinct()
        return Response(list(versions))
    
    @action(detail=False, methods=['get'])
    def get_login_methods(self, request):
        """根据设备类型、品牌(厂商)和版本获取认证方式"""
        device_type = request.query_params.get('device_type')
        manufacturer = request.query_params.get('manufacturer')
        version = request.query_params.get('version')
        if not all([device_type, manufacturer, version]):
            return Response({'error': 'device_type、manufacturer和version参数必填'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        login_methods = ManufacturerVersionInfo.objects.filter(
            device_type=device_type,
            manufacturer=manufacturer,
            version=version
        ).values_list('auth_method', flat=True).distinct()
        return Response(list(login_methods))


class ColumnOptionsViewSet(viewsets.ModelViewSet):
    """列可选值管理API"""
    queryset = ColumnOptions.objects.all()
    serializer_class = ColumnOptionsSerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def get_options_by_column(self, request):
        """根据列名获取可选值"""
        column_name = request.query_params.get('column_name')
        if not column_name:
            return Response({'error': 'column_name参数必填'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        options = ColumnOptions.objects.filter(
            column_name=column_name
        ).values_list('option_value', flat=True)
        return Response(list(options))


class UCMDeviceInventoryViewSet(viewsets.ModelViewSet):
    """UCM设备清单管理API"""
    queryset = UCMDeviceInventory.objects.all()
    serializer_class = UCMDeviceInventorySerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['post'])
    def upload_inventory(self, request):
        """上传UCM设备清单Excel"""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # 解析Excel文件
            workbook = xlrd.open_workbook(file_contents=file.read())
            sheet = workbook.sheet_by_index(0)
            
            # 获取列名（第一行）
            headers = [str(cell.value).strip() for cell in sheet.row(0)]
            
            # 清空现有数据
            UCMDeviceInventory.objects.all().delete()
            
            # 导入数据
            import_time = timezone.now()
            success_count = 0
            error_rows = []
            
            for row_idx in range(1, sheet.nrows):
                try:
                    row_data = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < sheet.ncols:
                            row_data[header] = str(sheet.cell(row_idx, col_idx).value).strip()
                    
                    # 创建设备记录
                    device = UCMDeviceInventory(
                        name=row_data.get('名称', ''),
                        device_type=row_data.get('设备类型', ''),
                        manufacturer=row_data.get('品牌(厂商)', ''),
                        version=row_data.get('版本', ''),
                        ip=row_data.get('IP', ''),
                        other_ips=row_data.get('其他IP', ''),
                        location=row_data.get('安装位置', ''),
                        group=row_data.get('分组', ''),
                        auth_method=row_data.get('认证方式', ''),
                        import_time=import_time
                    )
                    device.save()
                    success_count += 1
                    
                except Exception as e:
                    error_rows.append({
                        'row': row_idx + 1,
                        'error': str(e)
                    })
            
            return Response({
                'success': True,
                'message': f'成功导入 {success_count} 条记录',
                'errors': error_rows
            })
            
        except Exception as e:
            return Response({'error': f'文件解析失败: {str(e)}'}, 
                          status=status.HTTP_400_BAD_REQUEST)


class UCMRequirementViewSet(viewsets.ModelViewSet):
    """UCM需求登记管理API"""
    queryset = UCMRequirement.objects.all()
    serializer_class = UCMRequirementSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        # 获取筛选参数
        status_filter = self.request.query_params.get('status', 'pending')
        ucm_change_date = self.request.query_params.get('ucm_change_date')
        requirement_type = self.request.query_params.get('requirement_type')
        search = self.request.query_params.get('search')
        submitter = self.request.query_params.get('submitter')
        
        # 构建查询条件
        queryset = UCMRequirement.objects.all()
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if ucm_change_date:
            queryset = queryset.filter(ucm_change_date=ucm_change_date)
        if requirement_type:
            queryset = queryset.filter(requirement_type=requirement_type)
        if submitter:
            queryset = queryset.filter(submitter__username=submitter)
        if search:
            queryset = queryset.filter(
                Q(device_name__icontains=search) | Q(ip__icontains=search)
            )
        
        # 排序
        return queryset.order_by('-submit_time')
    
    @action(detail=False, methods=['post'])
    def upload_excel(self, request):
        """上传Excel文件并解析"""
        file = request.FILES.get('file')
        requirement_type = request.data.get('requirement_type')
        
        if not file:
            return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)
        if not requirement_type:
            return Response({'error': '请选择需求类型'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # 解析Excel
            workbook = xlrd.open_workbook(file_contents=file.read())
            sheet = workbook.sheet_by_index(0)
            
            # 获取列名和数据
            headers = [str(cell.value).strip() for cell in sheet.row(0)]
            data = []
            
            for row_idx in range(1, sheet.nrows):
                row_data = {}
                for col_idx, header in enumerate(headers):
                    if col_idx < sheet.ncols:
                        row_data[header] = str(sheet.cell(row_idx, col_idx).value).strip()
                data.append(row_data)
            
            return Response({
                'headers': headers,
                'data': data,
                'total_rows': len(data)
            })
            
        except Exception as e:
            return Response({'error': f'文件解析失败: {str(e)}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def available_dates(self, request):
        """获取可用的UCM变更日期"""
        try:
            from datetime import timedelta

            config = UCMDateConfig.objects.first()
            if not config:
                config = UCMDateConfig.objects.create(
                    wednesday_deadline_hours=7,
                    saturday_deadline_hours=31
                )

            now = timezone.now()
            available_dates = []
            deadlines = {}

            # 计算本周三
            this_wednesday = now + timedelta(days=(2 - now.weekday()) % 7)
            this_wednesday = this_wednesday.replace(hour=0, minute=0, second=0, microsecond=0)

            # 计算本周六
            this_saturday = now + timedelta(days=(5 - now.weekday()) % 7)
            this_saturday = this_saturday.replace(hour=0, minute=0, second=0, microsecond=0)

            # 生成本周及未来3周的周三和周六
            candidates = []
            for week in range(4):  # 0=本周, 1=下周, 2=3周后, 3=4周后
                # 添加周三
                wednesday = this_wednesday + timedelta(weeks=week)
                wednesday_dl = wednesday - timedelta(hours=config.wednesday_deadline_hours)
                candidates.append((wednesday, wednesday_dl, '周三'))

                # 添加周六
                saturday = this_saturday + timedelta(weeks=week)
                saturday_dl = saturday - timedelta(hours=config.saturday_deadline_hours)
                candidates.append((saturday, saturday_dl, '周六'))

            # 星期映射
            weekday_map = {
                0: '周一',
                1: '周二',
                2: '周三',
                3: '周四',
                4: '周五',
                5: '周六',
                6: '周日'
            }

            for date, deadline, day_type in candidates:
                if now < deadline:
                    date_str = date.strftime('%Y-%m-%d')
                    # 获取截止日期的星期几
                    deadline_weekday = weekday_map[deadline.weekday()]
                    deadline_str = f"{deadline.strftime('%Y-%m-%d')}{deadline_weekday} {deadline.strftime('%H:%M')}"
                    available_dates.append(date_str)
                    deadlines[date_str] = f"{day_type}UCM变更，最晚登记时间在{deadline_str}前"

            return Response({
                'dates': available_dates,
                'deadlines': deadlines
            })
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def list_dates(self, request):
        """获取需求列表页的可选日期（返回所有有数据的日期，不考虑截止时间）"""
        try:
            from django.db.models import Count
            from datetime import timedelta

            now = timezone.now()

            # 计算本周三和本周六
            this_wednesday = now + timedelta(days=(2 - now.weekday()) % 7)
            this_wednesday = this_wednesday.replace(hour=0, minute=0, second=0, microsecond=0)

            this_saturday = now + timedelta(days=(5 - now.weekday()) % 7)
            this_saturday = this_saturday.replace(hour=0, minute=0, second=0, microsecond=0)

            # 计算下周三和下周六
            next_wednesday = this_wednesday + timedelta(days=7)
            next_saturday = this_saturday + timedelta(days=7)

            # 生成所有可能的周三和周六日期（过去8周到未来8周）
            all_possible_dates = []
            for week in range(-8, 9):
                week_start = this_wednesday + timedelta(days=7 * week)
                wednesday = week_start
                saturday = week_start + timedelta(days=3)
                all_possible_dates.append(wednesday.strftime('%Y-%m-%d'))
                all_possible_dates.append(saturday.strftime('%Y-%m-%d'))

            # 查询数据库，获取有数据的日期
            requirements = UCMRequirement.objects.filter(
                ucm_change_date__in=all_possible_dates
            ).values('ucm_change_date').annotate(
                count=Count('id')
            ).order_by('ucm_change_date')

            # 提取有数据的日期
            list_dates = [req['ucm_change_date'] for req in requirements]

            return Response({
                'dates': list_dates
            })
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def validate_data(self, request):
        """校验数据"""
        requirement_type = request.data.get('requirement_type')
        excel_data = request.data.get('excel_data', [])
        
        if not requirement_type or not excel_data:
            return Response({'error': '参数不完整'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 获取模板配置
        try:
            template = TemplateConfig.objects.get(template_type=requirement_type)
            template_columns = template.get_column_definitions()
            template_column_names = [col['name'] for col in template_columns]
        except TemplateConfig.DoesNotExist:
            return Response({'error': '模板配置不存在'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 列名校验
        excel_headers = list(excel_data[0].keys()) if excel_data else []
        missing_columns = []
        extra_columns = []
        
        # 检查缺少的列
        for col_def in template_columns:
            if col_def['name'] not in excel_headers:
                missing_columns.append(col_def['name'])
        
        # 检查多余的列
        for header in excel_headers:
            if header not in template_column_names:
                extra_columns.append(header)
        
        if missing_columns or extra_columns:
            return Response({
                'valid': False,
                'error_type': 'column_mismatch',
                'missing_columns': missing_columns,
                'extra_columns': extra_columns
            })
        
        # 数据校验
        validation_results = []
        column_options = {}
        
        # 获取所有列的可选值
        for col in excel_headers:
            options = list(ColumnOptions.objects.filter(
                column_name=col
            ).values_list('option_value', flat=True))
            if options:
                column_options[col] = options
        
        for row_idx, row_data in enumerate(excel_data):
            row_validation = {
                'row_index': row_idx,
                'is_valid': True,
                'errors': {},
                'warnings': {}
            }
            
            # 校验每列数据
            for col_def in template_columns:
                col_name = col_def['name']
                value = row_data.get(col_name, '').strip()
                
                # 必填字段校验
                if col_def['required'] and not value:
                    row_validation['errors'][col_name] = '此字段为必填项'
                    row_validation['is_valid'] = False
                    continue
                
                # IP地址格式校验
                if col_name == 'IP' and value:
                    import re
                    ipv4_pattern = r'^(\d{1,3}\.){3}\d{1,3}$'
                    if not re.match(ipv4_pattern, value):
                        row_validation['errors'][col_name] = 'IP地址格式不正确（IPv4）'
                        row_validation['is_valid'] = False
                        continue
                
                # 可选值校验
                if col_name in column_options and value:
                    if value not in column_options[col_name]:
                        row_validation['errors'][col_name] = '不在可选值清单中'
                        row_validation['is_valid'] = False
                
                # 级联关系校验（设备类型→品牌(厂商)→版本）
            device_type = row_data.get('设备类型', '').strip()
            manufacturer = row_data.get('品牌(厂商)', '').strip()
            version = row_data.get('版本', '').strip()

            # 规则：如果选择了设备类型，必须选择品牌(厂商)
            if device_type and not manufacturer:
                row_validation['errors']['品牌(厂商)'] = '请选择品牌(厂商)'
                row_validation['is_valid'] = False

            # 规则：如果选择了品牌(厂商)，必须选择版本
            if manufacturer and not version:
                row_validation['errors']['版本'] = '请选择版本'
                row_validation['is_valid'] = False

            # 规则：如果三者都填了，检查组合是否有效
            if device_type and manufacturer and version:
                is_valid_combination = ManufacturerVersionInfo.objects.filter(
                    device_type=device_type,
                    manufacturer=manufacturer,
                    version=version
                ).exists()

                if not is_valid_combination:
                    row_validation['errors']['版本'] = '设备类型、品牌(厂商)、版本组合不匹配'
                    row_validation['is_valid'] = False

            validation_results.append(row_validation)
        
        return Response({
            'valid': True,
            'validation_results': validation_results
        })
    
    @action(detail=False, methods=['post'])
    def check_duplicates(self, request):
        """检查重复需求"""
        ucm_change_date = request.data.get('ucm_change_date')
        requirements = request.data.get('requirements', [])
        
        if not ucm_change_date or not requirements:
            return Response({'error': '参数不完整'}, status=status.HTTP_400_BAD_REQUEST)
        
        duplicates = []
        
        for req_data in requirements:
            name = req_data.get('名称', '')
            ip = req_data.get('IP', '')

            # 检查名称+UCM变更日期（仅当 name 不为空时检查）
            existing_by_name = None
            if name:
                existing_by_name = UCMRequirement.objects.filter(
                    device_name=name,
                    ucm_change_date=ucm_change_date,
                    status='pending'
                ).first()

            # 检查IP+UCM变更日期（仅当 ip 不为空时检查）
            existing_by_ip = None
            if ip:
                existing_by_ip = UCMRequirement.objects.filter(
                    ip=ip,
                    ucm_change_date=ucm_change_date,
                    status='pending'
                ).first()
            
            if existing_by_name or existing_by_ip:
                duplicate_info = {
                    'name': name,
                    'ip': ip
                }
                if existing_by_name:
                    duplicate_info['duplicate_by_name'] = {
                        'existing_date': existing_by_name.ucm_change_date.strftime('%Y-%m-%d'),
                        'existing_submitter': existing_by_name.submitter.username
                    }
                if existing_by_ip:
                    duplicate_info['duplicate_by_ip'] = {
                        'existing_date': existing_by_ip.ucm_change_date.strftime('%Y-%m-%d'),
                        'existing_submitter': existing_by_ip.submitter.username
                    }
                duplicates.append(duplicate_info)
        
        return Response({
            'has_duplicates': len(duplicates) > 0,
            'duplicates': duplicates
        })
    
    @action(detail=False, methods=['post'])
    def batch_submit(self, request):
        """批量提交需求（带事务）"""
        requirement_type = request.data.get('requirement_type')
        ucm_change_date = request.data.get('ucm_change_date')
        requirements = request.data.get('requirements', [])
        
        if not all([requirement_type, ucm_change_date, requirements]):
            return Response({'error': '参数不完整'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 使用事务确保数据一致性
        from django.db import transaction
        
        try:
            with transaction.atomic():
                submitted_count = 0
                skipped_count = 0
                skipped_records = []
                submitted_ids = []
                
                for req_data in requirements:
                    name = req_data.get('名称', '')
                    ip = req_data.get('IP', '')

                    # 检查重复（名称+UCM变更日期 或 IP+UCM变更日期）
                    # 仅当字段不为空时才进行重复检查
                    query = Q(ucm_change_date=ucm_change_date, status='pending')
                    if name:
                        query |= Q(device_name=name)
                    if ip:
                        query |= Q(ip=ip)

                    # 如果没有任何有效字段，跳过重复检查
                    if not name and not ip:
                        existing = None
                    else:
                        existing = UCMRequirement.objects.filter(query).first()
                    
                    if existing:
                        skipped_count += 1
                        skipped_records.append({
                            'name': name,
                            'ip': ip,
                            'reason': '重复记录'
                        })
                        continue
                    
                    # 创建需求记录
                    requirement = UCMRequirement(
                        requirement_type=requirement_type,
                        ucm_change_date=ucm_change_date,
                        submitter=request.user,
                        requirement_data=json.dumps(req_data, ensure_ascii=False),
                        device_name=name,
                        ip=ip
                    )
                    requirement.save()
                    submitted_count += 1
                    submitted_ids.append(requirement.id)
                
                return Response({
                    'success': True,
                    'submitted_count': submitted_count,
                    'skipped_count': skipped_count,
                    'skipped_records': skipped_records,
                    'submitted_ids': submitted_ids
                })
        except Exception as e:
            transaction.set_rollback(True)
            return Response({'error': f'提交失败: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def submit_requirement(self, request):
        """提交需求登记"""
        requirement_type = request.data.get('requirement_type')
        ucm_change_date = request.data.get('ucm_change_date')
        excel_data = request.data.get('excel_data', [])
        validation_results = request.data.get('validation_results', [])
        
        if not all([requirement_type, ucm_change_date, excel_data]):
            return Response({'error': '参数不完整'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 检查是否有校验失败的记录
        has_errors = any(
            len(result.get('errors', {})) > 0 
            for result in validation_results
        )
        if has_errors:
            return Response({'error': '存在校验失败的记录，无法提交'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # 检查重复记录
        duplicate_records = []
        for row_data in excel_data:
            name = row_data.get('名称', '')
            ip = row_data.get('IP', '')
            
            existing_requirement = UCMRequirement.objects.filter(
                Q(device_name=name) | Q(ip=ip),
                ucm_change_date=ucm_change_date,
                status='pending'
            ).first()
            
            if existing_requirement:
                duplicate_records.append({
                    'device_name': name,
                    'ip': ip,
                    'existing_date': existing_requirement.ucm_change_date
                })
        
        if duplicate_records:
            return Response({
                'error': '存在重复记录',
                'duplicate_records': duplicate_records
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # 创建需求记录
        success_count = 0
        for row_data in excel_data:
            try:
                requirement = UCMRequirement(
                    requirement_type=requirement_type,
                    ucm_change_date=ucm_change_date,
                    submitter=request.user,
                    requirement_data=json.dumps(row_data, ensure_ascii=False),
                    device_name=row_data.get('名称', ''),
                    ip=row_data.get('IP', '')
                )
                requirement.save()
                success_count += 1
            except Exception as e:
                continue
        
        return Response({
            'success': True,
            'message': f'成功登记 {success_count} 条需求'
        })
    
    @action(detail=True, methods=['post'])
    def mark_as_processed(self, request, pk=None):
        """标记需求为已处理"""
        requirement = self.get_object()
        requirement.status = 'processed'
        requirement.processor = request.user
        requirement.process_time = timezone.now()
        requirement.save()
        return Response({'success': True})
    
    @action(detail=False, methods=['post'])
    def batch_complete(self, request):
        """批量完成需求"""
        requirement_ids = request.data.get('requirement_ids', [])
        if not requirement_ids:
            return Response({'error': '请选择要完成的记录'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        count = UCMRequirement.objects.filter(
            id__in=requirement_ids,
            status='pending'
        ).update(
            status='processed',
            processor=request.user,
            process_time=timezone.now()
        )
        
        return Response({'success': True, 'count': count})
    
    @action(detail=False, methods=['post'])
    def batch_delete(self, request):
        """批量删除需求"""
        requirement_ids = request.data.get('requirement_ids', [])
        if not requirement_ids:
            return Response({'error': '请选择要删除的记录'},
                          status=status.HTTP_400_BAD_REQUEST)

        count, _ = UCMRequirement.objects.filter(
            id__in=requirement_ids
        ).delete()

        return Response({'success': True, 'count': count})

    @action(detail=False, methods=['get'])
    def weekly_dates(self, request):
        """获取指定周的周三、周六日期列表"""
        try:
            import math
            from datetime import timedelta
            from django.db.models import Min

            week_offset = int(request.query_params.get('week_offset', 0))
            now = timezone.now()

            # 计算目标周的周三和周六
            # 本周三是相对于当前日期的第几天 (周一=0, 周三=2, 周六=5)
            days_to_wednesday = (2 - now.weekday()) % 7
            days_to_saturday = (5 - now.weekday()) % 7

            # 加上周偏移
            target_wednesday = now + timedelta(days=days_to_wednesday + week_offset * 7)
            target_saturday = now + timedelta(days=days_to_saturday + week_offset * 7)

            # 获取周标签
            def get_week_label(offset):
                if offset == 0:
                    return '本周'
                elif offset == 1:
                    return '下周'
                elif offset == -1:
                    return '上周'
                else:
                    return f'{abs(offset)}周前' if offset < 0 else f'{offset}周后'

            week_label = get_week_label(week_offset)

            dates = [
                {
                    'date': target_wednesday.strftime('%Y-%m-%d'),
                    'day_type': '周三',
                    'label': f'{target_wednesday.strftime("%Y年%m月%d日")}（{week_label}三）'
                },
                {
                    'date': target_saturday.strftime('%Y-%m-%d'),
                    'day_type': '周六',
                    'label': f'{target_saturday.strftime("%Y年%m月%d日")}（{week_label}六）'
                }
            ]

            # 计算边界信息
            # 查找最早有数据的日期
            earliest_date_result = UCMRequirement.objects.aggregate(Min('ucm_change_date'))
            earliest_date = earliest_date_result['ucm_change_date__min']

            if earliest_date:
                # 将date对象转换为datetime对象，以便与now相减
                if not hasattr(earliest_date, 'hour'):
                    # earliest_date是date对象，需要转换为datetime
                    earliest_date = datetime.combine(earliest_date, time(0, 0, 0))
                # 计算最早日期相对于当前日期的周偏移
                days_diff = (earliest_date - now).days
                min_week_offset = math.floor(days_diff / 7)
            else:
                min_week_offset = 0

            # 最大week_offset固定为1（只允许到下周）
            max_week_offset = 1

            boundaries = {
                'min_week_offset': min_week_offset,
                'max_week_offset': max_week_offset
            }

            return Response({
                'dates': dates,
                'boundaries': boundaries
            })
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def date_statistics(self, request):
        """获取指定日期的需求类型统计"""
        try:
            date_str = request.query_params.get('date')
            if not date_str:
                return Response({'error': '请指定日期'},
                              status=status.HTTP_400_BAD_REQUEST)

            # 解析日期
            from datetime import datetime
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()

            # 统计各类型数量
            statistics = {
                'import': {'count': 0, 'pending': 0, 'processed': 0},
                'delete': {'count': 0, 'pending': 0, 'processed': 0},
                'modify': {'count': 0, 'pending': 0, 'processed': 0}
            }

            for req_type in ['import', 'delete', 'modify']:
                # 总数
                total_count = UCMRequirement.objects.filter(
                    ucm_change_date=target_date,
                    requirement_type=req_type
                ).count()

                # 待处理数
                pending_count = UCMRequirement.objects.filter(
                    ucm_change_date=target_date,
                    requirement_type=req_type,
                    status='pending'
                ).count()

                # 已处理数
                processed_count = UCMRequirement.objects.filter(
                    ucm_change_date=target_date,
                    requirement_type=req_type,
                    status='processed'
                ).count()

                statistics[req_type] = {
                    'count': total_count,
                    'pending': pending_count,
                    'processed': processed_count
                }

            return Response({
                'date': date_str,
                'statistics': statistics
            })
        except ValueError:
            return Response({'error': '日期格式错误，请使用 YYYY-MM-DD 格式'},
                          status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """导出需求列表为Excel文件"""
        try:
            # 获取筛选参数
            status_filter = request.query_params.get('status', 'pending')
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            requirement_type = request.query_params.get('requirement_type')
            
            # 构建查询条件
            queryset = UCMRequirement.objects.all()
            
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            if start_date:
                queryset = queryset.filter(ucm_change_date__gte=start_date)
            if end_date:
                queryset = queryset.filter(ucm_change_date__lte=end_date)
            if requirement_type:
                queryset = queryset.filter(requirement_type=requirement_type)
            
            # 排序
            queryset = queryset.order_by('-submit_time')
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "需求列表"
            
            # 定义样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 定义表头
            headers = [
                '需求ID',
                '需求类型',
                '名称',
                'IP地址',
                'UCM变更日期',
                '提交人',
                '提交时间',
                '状态',
                '处理人',
                '处理时间',
                '需求详情'
            ]
            
            # 写入表头
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
            
            # 写入数据
            row_idx = 2
            for requirement in queryset:
                # 解析需求数据
                try:
                    requirement_data = json.loads(requirement.requirement_data) if requirement.requirement_data else {}
                except:
                    requirement_data = {}
                
                # 写入行数据
                ws.cell(row=row_idx, column=1, value=requirement.id).border = border
                ws.cell(row=row_idx, column=2, value=requirement.get_requirement_type_display()).border = border
                ws.cell(row=row_idx, column=3, value=requirement.device_name).border = border
                ws.cell(row=row_idx, column=4, value=requirement.ip).border = border
                ws.cell(row=row_idx, column=5, value=requirement.ucm_change_date.strftime('%Y-%m-%d') if requirement.ucm_change_date else '').border = border
                ws.cell(row=row_idx, column=6, value=requirement.submitter.username if requirement.submitter else '').border = border
                ws.cell(row=row_idx, column=7, value=requirement.submit_time.strftime('%Y-%m-%d %H:%M:%S') if requirement.submit_time else '').border = border
                ws.cell(row=row_idx, column=8, value=requirement.get_status_display()).border = border
                ws.cell(row=row_idx, column=9, value=requirement.processor.username if requirement.processor else '').border = border
                ws.cell(row=row_idx, column=10, value=requirement.process_time.strftime('%Y-%m-%d %H:%M:%S') if requirement.process_time else '').border = border
                
                # 需求详情（将JSON转换为可读格式）
                detail_text = ""
                if requirement_data:
                    detail_items = []
                    for key, value in requirement_data.items():
                        if key not in ['名称', 'IP']:  # 避免重复
                            detail_items.append(f"{key}: {value}")
                    detail_text = "; ".join(detail_items)
                
                detail_cell = ws.cell(row=row_idx, column=11, value=detail_text)
                detail_cell.border = border
                detail_cell.alignment = Alignment(wrap_text=True)
                
                row_idx += 1
            
            # 调整列宽
            column_widths = [10, 15, 20, 15, 15, 15, 20, 10, 15, 20, 50]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width
            
            # 创建HTTP响应
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            # 设置文件名
            status_text = '待处理' if status_filter == 'pending' else '已处理'
            filename = f'UCM需求列表_{status_text}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            # 保存工作簿到响应
            wb.save(response)

            return response

        except Exception as e:
            return Response({'error': f'导出失败: {str(e)}'},
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def export_change_plan(self, request):
        """导出变更方案"""
        import traceback
        ucm_change_date = request.data.get('ucm_change_date')

        if not ucm_change_date:
            return Response({'error': '请提供UCM变更日期'},
                          status=status.HTTP_400_BAD_REQUEST)

        try:
            # 查询选中日期的所有需求数据
            requirements = UCMRequirement.objects.filter(
                ucm_change_date=ucm_change_date
            )

            print(f"查询到 {requirements.count()} 条需求数据")

            # 按类型和地点分组
            grouped_data = self._group_by_type_and_location(requirements)

            print(f"分组数据: {grouped_data}")

            # 生成所有文件
            files = []

            # 生成变更方案
            change_files = self._generate_change_plans(grouped_data, ucm_change_date)
            files.extend(change_files)
            print(f"生成了 {len(change_files)} 个变更方案文件")

            # 生成回退方案
            rollback_files = self._generate_rollback_plans(grouped_data, ucm_change_date)
            files.extend(rollback_files)
            print(f"生成了 {len(rollback_files)} 个回退方案文件")

            # 如果没有文件，返回提示
            if not files:
                return Response({'error': '该日期没有需求数据'},
                              status=status.HTTP_400_BAD_REQUEST)

            # 创建压缩包
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for file_info in files:
                    zip_file.writestr(file_info['filename'], file_info['data'])

            zip_buffer.seek(0)

            # 返回压缩包
            response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="UCM变更方案_{ucm_change_date}.zip"'

            return response

        except Exception as e:
            error_msg = f'导出失败: {str(e)}\n{traceback.format_exc()}'
            print(error_msg)
            return Response({'error': error_msg},
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _group_by_type_and_location(self, requirements):
        """按类型和地点分组"""
        grouped = {
            'import': {'外高桥': [], '嘉定': [], '境外机构': [], '分行': []},
            'modify': {'外高桥': [], '嘉定': [], '境外机构': [], '分行': []},
            'delete': {'外高桥': [], '嘉定': [], '境外机构': [], '分行': []},
        }

        for req in requirements:
            req_type = req.requirement_type
            req_data = json.loads(req.requirement_data) if req.requirement_data else {}

            # 获取地点
            location = self._get_location(req_data, req_type)

            # 添加到对应分组
            if location in grouped[req_type]:
                grouped[req_type][location].append(req_data)

        return grouped

    def _get_location(self, row, requirement_type):
        """根据IP和名称识别地点"""
        ip_column = 'IP' if requirement_type in ['import', 'delete'] else '设备ip'
        name_column = '名称' if requirement_type in ['import', 'delete'] else '老指标'

        ip = row.get(ip_column, '')
        name = row.get(name_column, '')

        # 优先根据IP前缀判断
        if ip.startswith('76.'):
            return '嘉定'
        elif ip.startswith('84.'):
            return '外高桥'
        elif ip.startswith('123.'):
            return '境外机构'

        # 根据名称前2位判断
        if name:
            prefix = name[:2].upper()
            if prefix == 'NF':
                return '外高桥'
            elif prefix == 'JD':
                return '嘉定'
            elif prefix.isalpha():
                return '分行'

        # 默认返回分行
        return '分行'

    def _generate_change_plans(self, grouped_data, ucm_change_date):
        """生成变更方案文件"""
        files = []

        for req_type in ['import', 'modify', 'delete']:
            for location in ['外高桥', '嘉定', '境外机构', '分行']:
                data = grouped_data.get(req_type, {}).get(location, [])

                if not data:
                    continue

                # 根据类型生成文件
                if req_type == 'delete':
                    file_data = self._generate_delete_change_plan(data)
                    template_type = 'delete'
                elif req_type == 'modify':
                    file_data = self._generate_modify_change_plan(data)
                    template_type = 'modify'
                else:  # import
                    file_data = self._generate_import_change_plan(data)
                    template_type = 'import'

                type_name = {'import': '导入', 'modify': '修改', 'delete': '删除'}[req_type]
                filename = f'UCM_{type_name}_{location}.xls'

                files.append({'filename': filename, 'data': file_data})

        return files

    def _generate_rollback_plans(self, grouped_data, ucm_change_date):
        """生成回退方案文件"""
        files = []

        for req_type in ['import', 'modify', 'delete']:
            for location in ['外高桥', '嘉定', '境外机构', '分行']:
                data = grouped_data.get(req_type, {}).get(location, [])

                if not data:
                    continue

                # 根据类型生成文件
                if req_type == 'delete':
                    # 删除类型的回退方案：使用导入模板的列结构，数据按导入规则映射
                    file_data = self._generate_import_change_plan(data, is_rollback=True)
                    template_type = 'import'
                elif req_type == 'modify':
                    file_data = self._generate_modify_rollback_plan(data)
                    template_type = 'modify'
                else:  # import
                    file_data = self._generate_import_change_plan(data, is_rollback=True)
                    template_type = 'import'

                type_name = {'import': '导入', 'modify': '修改', 'delete': '删除'}[req_type]
                filename = f'UCM_{type_name}_{location}_回退.xls'

                files.append({'filename': filename, 'data': file_data})

        return files

    def _generate_delete_change_plan(self, data):
        """生成删除类型的变更方案"""
        try:
            # 获取删除模板
            template = TemplateConfig.objects.get(template_type='delete')
            columns = template.get_column_definitions()
        except TemplateConfig.DoesNotExist:
            columns = []

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "变更方案"

        # 删除模板的列名映射
        column_mapping = {
            '设备IP': 'IP',
            '需求': '删除',
            '老指标': '名称',
            '新指标': ''
        }

        # 添加表头
        headers = [col['name'] for col in columns]
        ws.append(headers)

        # 设置表头样式（通过列索引访问）
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="0050B3")
            cell.fill = PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # 添加数据行
        for row_data in data:
            row_values = []
            for col in columns:
                col_name = col['name']
                if col_name in column_mapping:
                    value = column_mapping[col_name]
                    if value == '删除':
                        row_values.append('删除')
                    elif value == '':
                        row_values.append('')
                    else:
                        row_values.append(row_data.get(value, ''))
                else:
                    row_values.append(row_data.get(col_name, ''))

            ws.append(row_values)

        # 保存到BytesIO
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _generate_modify_change_plan(self, data):
        """生成修改类型的变更方案"""
        try:
            # 获取修改模板
            template = TemplateConfig.objects.get(template_type='modify')
            columns = template.get_column_definitions()
        except TemplateConfig.DoesNotExist:
            columns = []

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "变更方案"

        # 添加表头
        headers = [col['name'] for col in columns]
        ws.append(headers)

        # 设置表头样式（通过列索引访问）
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="D46B08")
            cell.fill = PatternFill(start_color="FFF7E6", end_color="FFF7E6", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # 添加数据行（列名与需求列表一一对应）
        for row_data in data:
            row_values = [row_data.get(col['name'], '') for col in columns]
            ws.append(row_values)

        # 保存到BytesIO
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _generate_modify_rollback_plan(self, data):
        """生成修改类型的回退方案"""
        try:
            # 获取修改模板
            template = TemplateConfig.objects.get(template_type='modify')
            columns = template.get_column_definitions()
        except TemplateConfig.DoesNotExist:
            columns = []

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "回退方案"

        # 添加表头
        headers = [col['name'] for col in columns]
        ws.append(headers)

        # 设置表头样式（通过列索引访问）
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="D46B08")
            cell.fill = PatternFill(start_color="FFF7E6", end_color="FFF7E6", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # 添加数据行（老指标和新指标互换）
        for row_data in data:
            row_values = []
            for col in columns:
                col_name = col['name']
                if col_name == '老指标':
                    row_values.append(row_data.get('新指标', ''))
                elif col_name == '新指标':
                    row_values.append(row_data.get('老指标', ''))
                else:
                    row_values.append(row_data.get(col_name, ''))
            ws.append(row_values)

        # 保存到BytesIO
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _generate_import_change_plan(self, data, is_rollback=False):
        """生成导入类型的变更方案或回退方案"""
        try:
            # 获取导入模板
            template = TemplateConfig.objects.get(template_type='import')
            columns = template.get_column_definitions()
        except TemplateConfig.DoesNotExist:
            columns = []

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "变更方案" if not is_rollback else "回退方案"

        # 添加表头
        headers = [col['name'] for col in columns]
        ws.append(headers)

        # 设置表头样式（通过列索引访问）
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="0050B3")
            cell.fill = PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # 添加数据行
        for row_data in data:
            if is_rollback:
                # 回退方案：需求→'删除'
                row_values = []
                for col in columns:
                    col_name = col['name']
                    if col_name == '需求':
                        row_values.append('删除')
                    elif col_name == '新指标':
                        row_values.append('')
                    else:
                        row_values.append(row_data.get(col_name, ''))
                ws.append(row_values)
            else:
                # 变更方案：列名与需求列表一一对应
                row_values = [row_data.get(col['name'], '') for col in columns]
                ws.append(row_values)

        # 保存到BytesIO
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


class TemplateConfigViewSet(viewsets.ModelViewSet):
    """模板配置管理API"""
    queryset = TemplateConfig.objects.all()
    serializer_class = TemplateConfigSerializer
    permission_classes = [IsAuthenticated]
    
    def update(self, request, *args, **kwargs):
        """更新模板配置，验证列定义格式"""
        column_definitions = request.data.get('column_definitions')
        
        if column_definitions:
            try:
                # 尝试解析JSON
                if isinstance(column_definitions, str):
                    columns = json.loads(column_definitions)
                else:
                    columns = column_definitions
                
                # 验证格式
                if not isinstance(columns, list):
                    return Response(
                        {'error': 'column_definitions必须是数组'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # 验证每个列定义的格式
                for idx, col in enumerate(columns):
                    if not isinstance(col, dict):
                        return Response(
                            {'error': f'列定义[{idx}]必须是对象'}, 
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    
                    if 'name' not in col or not isinstance(col['name'], str):
                        return Response(
                            {'error': f'列定义[{idx}]必须包含name字段（字符串）'}, 
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    
                    if 'required' not in col or not isinstance(col['required'], bool):
                        return Response(
                            {'error': f'列定义[{idx}]必须包含required字段（布尔值）'}, 
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    
                    if 'example' not in col or not isinstance(col['example'], str):
                        return Response(
                            {'error': f'列定义[{idx}]必须包含example字段（字符串）'}, 
                            status=status.HTTP_400_BAD_REQUEST
                        )
                
                # 确保数据是JSON字符串格式
                request.data['column_definitions'] = json.dumps(columns, ensure_ascii=False)
                
            except json.JSONDecodeError:
                return Response(
                    {'error': 'column_definitions必须是有效的JSON'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        return super().update(request, *args, **kwargs)
    
    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """下载模板文件（.xls格式）"""
        template_type = request.query_params.get('template_type')

        if not template_type:
            return Response({'error': '请指定模板类型'},
                          status=status.HTTP_400_BAD_REQUEST)

        # 获取模板配置
        try:
            template = TemplateConfig.objects.get(template_type=template_type)
            columns = template.get_column_definitions()
        except TemplateConfig.DoesNotExist:
            return Response({'error': '模板配置不存在'},
                          status=status.HTTP_400_BAD_REQUEST)

        # 使用xlwt创建.xls文件
        import xlwt

        # 创建工作簿
        workbook = xlwt.Workbook(encoding='utf-8')
        sheet = workbook.add_sheet('模板')

        # 定义样式
        header_style = xlwt.XFStyle()
        header_font = xlwt.Font()
        header_font.bold = True
        header_font.height = 280  # 14pt
        header_style.font = header_font

        # 根据模板类型设置不同的表头颜色
        color_map = {
            'import': 'light_blue',    # 浅蓝色
            'modify': 'light_orange',  # 浅橙色
            'delete': 'light_red'      # 浅红色
        }
        header_color = color_map.get(template_type, 'light_blue')

        header_pattern = xlwt.Pattern()
        header_pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        header_pattern.pattern_fore_colour = xlwt.Style.colour_map[header_color]
        header_style.pattern = header_pattern

        header_alignment = xlwt.Alignment()
        header_alignment.horz = xlwt.Alignment.HORZ_CENTER
        header_alignment.vert = xlwt.Alignment.VERT_CENTER
        header_style.alignment = header_alignment

        header_borders = xlwt.Borders()
        header_borders.left = xlwt.Borders.THIN
        header_borders.right = xlwt.Borders.THIN
        header_borders.top = xlwt.Borders.THIN
        header_borders.bottom = xlwt.Borders.THIN
        header_style.borders = header_borders

        data_style = xlwt.XFStyle()
        data_borders = xlwt.Borders()
        data_borders.left = xlwt.Borders.THIN
        data_borders.right = xlwt.Borders.THIN
        data_borders.top = xlwt.Borders.THIN
        data_borders.bottom = xlwt.Borders.THIN
        data_style.borders = data_borders

        # 写入表头
        for col_idx, col_def in enumerate(columns):
            header = col_def['name']
            sheet.write(0, col_idx, header, header_style)
            # 设置列宽
            sheet.col(col_idx).width = 4000  # 约20个字符

        # 写入样例数据
        for col_idx, col_def in enumerate(columns):
            if col_def['example']:
                sheet.write(1, col_idx, col_def['example'], data_style)

        # 创建HTTP响应
        response = HttpResponse(content_type='application/vnd.ms-excel')

        # 设置文件名
        type_map = {
            'import': '导入',
            'modify': '修改',
            'delete': '删除'
        }
        type_name = type_map.get(template_type, template_type)
        filename = f'{type_name}_模板.xls'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # 保存工作簿到响应
        workbook.save(response)

        return response


@api_view(['POST'])
@permission_classes([AllowAny])
def user_login(request):
    """用户登录API"""
    username = request.data.get('username')
    password = request.data.get('password')
    
    if not username or not password:
        return Response({'error': '请输入用户名和密码'}, 
                      status=status.HTTP_400_BAD_REQUEST)
    
    user = authenticate(username=username, password=password)
    if user is not None:
        login(request, user)
        return Response({
            'success': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'is_staff': user.is_staff
            }
        })
    else:
        return Response({'error': '用户名或密码错误'}, 
                      status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def user_logout(request):
    """用户退出API"""
    logout(request)
    return Response({'success': True})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    """修改密码API"""
    try:
        old_password = request.data.get('old_password')
        new_password = request.data.get('new_password')
        
        if not old_password or not new_password:
            return Response({'error': '请提供原密码和新密码'}, status=status.HTTP_400_BAD_REQUEST)
        
        if len(new_password) < 6:
            return Response({'error': '新密码至少6位'}, status=status.HTTP_400_BAD_REQUEST)
        
        user = request.user
        
        # 验证原密码
        if not user.check_password(old_password):
            return Response({'error': '原密码错误'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 修改密码
        user.set_password(new_password)
        user.save()
        
        return Response({'success': True, 'message': '密码修改成功'})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_current_user(request):
    """获取当前登录用户信息"""
    user = request.user
    return Response({
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'is_staff': user.is_staff
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_ucm_date_config(request):
    """获取UCM日期配置"""
    try:
        config = UCMDateConfig.objects.first()
        if not config:
            # 如果没有配置，创建默认配置
            config = UCMDateConfig.objects.create(
                wednesday_deadline_hours=7,
                saturday_deadline_hours=31
            )
        
        return Response({
            'wednesday_deadline_hours': config.wednesday_deadline_hours,
            'saturday_deadline_hours': config.saturday_deadline_hours
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def deadline_config(request):
    """获取或更新登记截止配置"""
    try:
        config = UCMDateConfig.objects.first()
        if not config:
            config = UCMDateConfig.objects.create(
                wednesday_deadline_hours=7,
                saturday_deadline_hours=31
            )
        
        if request.method == 'GET':
            return Response({
                'wednesday_deadline_hours': config.wednesday_deadline_hours,
                'saturday_deadline_hours': config.saturday_deadline_hours
            })
        elif request.method == 'PUT':
            wednesday_hours = request.data.get('wednesday_deadline_hours')
            saturday_hours = request.data.get('saturday_deadline_hours')
            
            # 验证输入
            if wednesday_hours is not None:
                try:
                    wednesday_hours = int(wednesday_hours)
                    if not (-168 <= wednesday_hours <= 168):
                        return Response({'error': '周三截止小时数必须在-168到168之间'}, status=status.HTTP_400_BAD_REQUEST)
                except (ValueError, TypeError):
                    return Response({'error': '周三截止小时数格式错误'}, status=status.HTTP_400_BAD_REQUEST)
            
            if saturday_hours is not None:
                try:
                    saturday_hours = int(saturday_hours)
                    if not (-168 <= saturday_hours <= 168):
                        return Response({'error': '周六截止小时数必须在-168到168之间'}, status=status.HTTP_400_BAD_REQUEST)
                except (ValueError, TypeError):
                    return Response({'error': '周六截止小时数格式错误'}, status=status.HTTP_400_BAD_REQUEST)
            
            # 更新配置
            if wednesday_hours is not None:
                config.wednesday_deadline_hours = wednesday_hours
            if saturday_hours is not None:
                config.saturday_deadline_hours = saturday_hours
            
            config.save()
            
            return Response({
                'wednesday_deadline_hours': config.wednesday_deadline_hours,
                'saturday_deadline_hours': config.saturday_deadline_hours,
                'message': '配置更新成功'
            })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)